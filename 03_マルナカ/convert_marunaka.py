import argparse
import sys
import os
import warnings
import json
import re
from datetime import datetime

import constant

# These are libraries related to reading and editing .xlsx file.
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from xlcalculator import Evaluator
from xlcalculator import ModelCompiler


def extract_data(source_filename: str, verbose: bool = False) -> (list, dict):
    if verbose:
        print('Extracting data...')

    if not source_filename:
        sys.exit('Source must be provided by using "--source" or "-s".')
    elif not os.path.isfile(source_filename):
        sys.exit('Source must be a file.')

    # Get all products
    # If a product has no product ID, it will ignore it.
    if verbose:
        print('Extracting products...')

    products = []

    source = _Source(source_filename)
    ignored_rows = []
    for row in range(constant.SOURCE.ORDER_STARTING_ROW, source.max_row + 1):
        # Ignore products with no product ID
        if not source.get(constant.SOURCE.PRODUCT_ID_COLUMN, row):
            ignored_rows.append(row)
            continue
        product_id = source.get(constant.SOURCE.PRODUCT_ID_COLUMN, row)
        lot = source.get(constant.SOURCE.LOT_COLUMN, row)
        spec = source.get(constant.SOURCE.SPEC_COLUMN, row)

        products.append({'productId': product_id,
                         'lot': lot,
                         'spec': spec})
    warnings.warn('Row(s) ' + __condense_number_list(ignored_rows) + ' was ignored due to empty product ID.')

    if verbose:
        print('Products extraction finished')

    if verbose:
        print('Extracting Store orders...')

    store_orders = {}
    order_ending_column = source.max_column + 1
    for column in range(constant.SOURCE.ORDER_STARTING_COLUMN, source.max_column + 1):
        # Break if there's no value in the date cell
        if not source.get(column, constant.SOURCE.ORDER_DATE_ROW):
            break
        # Ignore columns with no store name
        if not source.get(column, constant.SOURCE.STORE_NAME_ROW):
            continue
        store_name = source.get(column, constant.SOURCE.STORE_NAME_ROW)
        order_date = source.get(column, constant.SOURCE.ORDER_DATE_ROW).strftime('%m/%d')

        # If the store doesn't appear before, create a new dict for it
        if not store_orders.get(store_name):
            store_orders[store_name] = {}
            store_orders[store_name]['order_dates'] = []
            store_orders[store_name]['quantities'] = []
            for _ in range(len(products)):
                store_orders[store_name]['quantities'].append([])
        # If the combination of the store name and order date appears before, break it
        if store_orders.get(store_name) and order_date in store_orders[store_name]['order_dates']:
            order_ending_column = column
            break
        store_orders[store_name]['order_dates'].append(order_date)

    product_index = 0
    for row in range(constant.SOURCE.ORDER_STARTING_ROW, source.max_row + 1):
        if row in ignored_rows:
            continue

        for column in range(constant.SOURCE.ORDER_STARTING_COLUMN, order_ending_column):
            order_date = source.get(column, constant.SOURCE.ORDER_DATE_ROW)
            store_name = source.get(column, constant.SOURCE.STORE_NAME_ROW)
            # Break if there's no value in the date cell, which means no order date
            if not order_date:
                break
            # Ignore columns with no store name
            if not store_name:
                continue
            # If there's no value, give it 0
            if source.get(column, row):
                store_orders[store_name]['quantities'][product_index].append(__extract_number(source.get(column, row)))
            else:
                store_orders[store_name]['quantities'][product_index].append(0)

        product_index += 1

    if verbose:
        print('Store orders extration finished.')

    if verbose:
        print('Data: -----------------------------------------------------')
        print(json.dumps(store_orders, indent=4))
        print('-----------------------------------------------------------')
        print('Data extraction finished.')

    return (products, store_orders)


def output_to_destination(products: list, store_orders: dict, destination: str, verbose: bool = False):
    if verbose:
        print('Output started.')

    if os.path.isfile(destination):
        sys.exit('Destination must be a directory.')
    os.makedirs(destination, exist_ok=True)

    # Put those datas in a new file with another format
    for store_name, store_order in store_orders.items():
        store_order_workbook = load_workbook(constant.TEMPLATE_FILENAME)
        store_order_worksheet = store_order_workbook.worksheets[0]
        store_order_worksheet.title = store_name

        for index, order_date in enumerate(store_order['order_dates']):
            order_date_cell = store_order_worksheet.cell(column=constant.DESTINATION.ORDER_STARTING_COLUMN + index,
                                                         row=constant.DESTINATION.ORDER_DATE_ROW)
            order_date_cell.value = order_date

        for product_index, product in enumerate(products):
            current_row = constant.DESTINATION.ORDER_STARTING_ROW + product_index
            product_id_cell = store_order_worksheet.cell(column=constant.DESTINATION.PRODUCT_ID_COLUMN,
                                                         row=current_row)
            lot_cell = store_order_worksheet.cell(column=constant.DESTINATION.LOT_COLUMN,
                                                  row=current_row)
            spec_cell = store_order_worksheet.cell(column=constant.DESTINATION.SPEC_COLUMN,
                                                   row=current_row)

            product_id_cell.value = product['productId']
            lot_cell.value = product['lot']
            spec_cell.value = product['spec']

            for quantity_index, quantity in enumerate(store_order['quantities'][product_index]):
                quantity_cell = store_order_worksheet.cell(column=constant.DESTINATION.ORDER_STARTING_COLUMN + quantity_index,
                                                           row=current_row)
                quantity_cell.value = quantity

        store_order_filename = os.path.join(destination, store_name + datetime.now().strftime('-%Y%m%dT%H%M%S') + '.xlsx')
        store_order_workbook.save(store_order_filename)

    if verbose:
        print('Output finished.')


def main():
    # Get arguments
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', '--source',
                        type=str,
                        default=None,
                        help='the filename of the source file')
    parser.add_argument('-d', '--destination',
                        type=str,
                        default='./results',
                        help='the destination of outputed files')
    parser.add_argument('-v', '--verbose',
                        action='store_true',
                        help='Show helpful logs')

    args = parser.parse_args()
    source_filename = args.source
    destination = args.destination
    verbose = args.verbose

    products, store_orders = extract_data(source_filename, verbose)
    output_to_destination(products, store_orders, destination, verbose)


class _Source():
    def __init__(self, source):
        source_workbook = load_workbook(source)
        if len(source_workbook.sheetnames) != 1:
            warnings.warn('Multiple worksheets found. Use the first one to continue.')
        self.__source_sheetname = source_workbook.sheetnames[0]
        source_sheet = source_workbook[self.__source_sheetname]
        self.max_column = source_sheet.max_column
        self.max_row = source_sheet.max_row

        source_compiler = ModelCompiler()
        source_model = source_compiler.read_and_parse_archive(source, build_code=True)
        self.__source_evaluator = Evaluator(source_model)

    def get(self, column_number, row_number):
        return self.__source_evaluator.get_cell_value(self.__source_sheetname
                                                      + '!'
                                                      + get_column_letter(column_number)
                                                      + str(row_number))


def __extract_number(text: str) -> int:
    return int(re.split(r'\D+', str(text))[-1])


def __condense_number_list(numbers: list) -> str:
    if len(numbers) == 0:
        return ''

    numbers = sorted(numbers)
    result = str(numbers[0])
    last_number = numbers[0]
    last_appended_number = numbers[0]
    for number in numbers[1:-1]:
        if last_number + 1 != number:
            if last_appended_number == last_number:
                result += f', {str(number)}'
            else:
                result += f'~{str(last_number)}, {str(number)}'
            last_appended_number = number
        last_number = number

    if last_number + 1 != numbers[-1]:
        result += f'~{str(last_number)}, {str(numbers[-1])}'
    else:
        result += f'~{str(numbers[-1])}'

    return f'[{result}]'


if __name__ == '__main__':
    main()
